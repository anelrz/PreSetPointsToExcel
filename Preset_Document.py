import openpyxl

class Preset_Document(object):

    def __init__(self, M_IDs):
        self.M_IDs = M_IDs
        
        
    def create(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "S-ID"    
        ws['B1'] = "Measurement" 
        ws['C1'] = "type"    
        ws['D1'] = "M-ID"    
        ws['E1'] = "Test gen"    
        ws['F1'] = "Setup"    
        ws['G1'] = "Test signal"    
        ws['H1'] = "Test level"    
        ws['I1'] = "Accept"    
        ws['J1'] = "Supply"    
        ws['K1'] = "Accept"
        i = 2
        
        for M_ID in self.M_IDs:
            for d in range(0, M_ID.testsignals.__len__()):
                ws.cell(row=i, column=4).value = M_ID.name + str(":" + str(d+1))
                ws.cell(row=i, column=7).value = str("%.2f " % (0 if M_ID.testsignals[d] == None else M_ID.testsignals[d])) + str(M_ID.testlevelunit)
                ws.cell(row=i, column=8).value = str("%.2f " % M_ID.testlevels[d]) + str(M_ID.unit)           
                i+=1
            i+=1

                    
        
        
        wb.save('editxlsx/' + "Signal Sheet.xlsx")
        
    def convert(self, value):
        pass
