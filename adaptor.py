import os
import openpyxl
#from src import M_ID_Collection
import M_ID_Collection
import Preset_Document
import sensorgroup
import string

if __name__ == '__main__':
    
    double_ratio = 0.8
    triple_mid_ratio = 0.6
    edge_multiply = 3
    double_edge_multiply = 2
    pre_unit = "0"
    prefix = input("Enter Prefix: ")
                
    #------Start---------------------------------------------------------   
    cou = 0
    cou1 = 0

    sg = sensorgroup.sensorgroup()
    

    collection = M_ID_Collection.M_ID_Collection(double_ratio, triple_mid_ratio, edge_multiply, double_edge_multiply, prefix)
    for file in os.listdir("myexcel"):
        if file.endswith(".xlsx"):
            wb = openpyxl.load_workbook('myexcel/' + file)
            for ws in wb.get_sheet_names():
                if(ws != 'Version'):
                    print('.',end="",flush=True)
                    cur_ws = wb.get_sheet_by_name(ws)

                    for row in range(2, cur_ws.get_highest_row()): 
                        if((cur_ws.cell(row=row, column=1).value != None) or (cur_ws.cell(row=row, column=6).value != None)  or (cur_ws.cell(row=row, column=8).value != None)):
                            cou1 +=1
    
                            if(cur_ws.cell(row=row, column=8).value == None): 
                                Hi = ""
                                myhi = None
                            elif(cur_ws.cell(row=row, column=8).value == "-"):
                                Hi = ""
                                myhi = None
                            else:
                                Hi = "', Hi='" + str(cur_ws.cell(row=row, column=8).value).replace(',', '.')
                                myhi = float(str(cur_ws.cell(row=row, column=8).value).replace(',', '.'))
                                
                            if(cur_ws.cell(row=row, column=9).value == None): 
                                HiHi = ""
                                myhihi = None
                            elif(cur_ws.cell(row=row, column=9).value == "-"):
                                HiHi = ""
                                myhihi = None
                            else:
                                HiHi = "', HiHi='" + str(cur_ws.cell(row=row, column=9).value).replace(',', '.')
                                myhihi = float(str(cur_ws.cell(row=row, column=9).value).replace(',', '.'))

                                
                    #---------------------------------------------------------------

                                
                            unit = cur_ws.cell(row=row, column=4).value
                            sensorgroup = sg.getgroup(str(cur_ws.cell(row=row, column=2).value))
                            
                            
                            if(unit == None):
                                unit = pre_unit
                                sensorgroup = pre_sensorgroup
                                
                            MID = collection.findandinsertMID(myhi, myhihi, unit, sensorgroup)
                            
                            if(MID != None):
                                cou += 1
                                col1 = MID.name + ':1'
                                col2 = MID.name + ':2'
                                
                                if(MID.Rhl == None and MID.Rhh == None):
                                    col3 = '-'
                                else:
                                    col3 = MID.name + ':3'
                                    
                                cur_ws.cell(row=row, column=18).value = col1
                                cur_ws.cell(row=row, column=29).value = col2
                                cur_ws.cell(row=row, column=35).value = col3
                            pre_unit = unit
                            pre_sensorgroup = sensorgroup
            print(' Saving.',end="",flush=True)
            wb.save('editxlsx/' + file)
            print (" Done")                  
    points = collection.M_ID_Choose_Points()
    PreDoc = Preset_Document.Preset_Document(points)
    PreDoc.create()
    
    
    print()
    print()
    for point in points:
        point.print_point()
    
    print()
    print("Lines set: " + str(cou) + " out of: " + str(cou1))
    
    print("Finished")
    input('')
    
